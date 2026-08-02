import base64
import io
import logging

from odoo import api, fields, models

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class WarehouseMapImportWizard(models.TransientModel):
    """Asistente para importar ubicaciones del mapa desde un archivo Excel.

    El Excel debe contener columnas identificables por nombre:
    rack / posicion / nivel / puesto (mayúsculas o minúsculas).
    Ejemplo: 'rack' = RACK A, 'posicion' = 1, 'nivel' = 1, 'puesto' = 1.
    """
    _name = 'warehouse.map.import.wizard'
    _description = 'Asistente de importación de ubicaciones desde Excel'

    file = fields.Binary(string="Archivo Excel", required=True)
    filename = fields.Char(string="Nombre del archivo")
    warehouse_id = fields.Many2one('stock.warehouse', string="Almacén", required=True)
    state = fields.Selection([
        ('choose', 'choose'),
        ('result', 'result'),
    ], default='choose')
    log = fields.Text(string="Resultado")

    def action_import(self):
        """Lee el Excel y crea/actualiza las ubicaciones del mapa."""
        self.ensure_one()
        log_lines = []
        success = 0
        errors = 0

        if openpyxl is None:
            self.log = "Error: La librería 'openpyxl' no está instalada. Ejecute: pip install openpyxl"
            self.state = 'result'
            return self._reload()

        try:
            data = io.BytesIO(base64.b64decode(self.file))
            wb = openpyxl.load_workbook(data, read_only=True)
            ws = wb.active

            # Mapea cabeceras del Excel a campos (por coincidencia parcial).
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_map = {
                'rack': None, 'posicion': None, 'nivel': None,
                'puesto': None, 'nombre': None, 'codigo': None,
            }
            for i, h in enumerate(headers):
                hl = (h or '').lower().strip()
                for key in col_map:
                    if key in hl:
                        col_map[key] = i

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    rack_val = row[col_map['rack']] if col_map['rack'] is not None else None
                    pos_val = row[col_map['posicion']] if col_map['posicion'] is not None else None
                    nivel_val = row[col_map['nivel']] if col_map['nivel'] is not None else None
                    puesto_val = row[col_map['puesto']] if col_map['puesto'] is not None else None

                    if not rack_val and not pos_val:
                        continue

                    rack_str = str(rack_val).strip().upper().replace('RACK ', '').replace('PASILLO ', 'P') if rack_val else ''
                    pos_int = int(float(pos_val)) if pos_val else 0
                    nivel_str = str(int(float(nivel_val))) if nivel_val else '1'
                    puesto_str = str(int(float(puesto_val))) if puesto_val else '1'

                    # Nombre canónico de la ubicación, p.ej. 'A01-N1-P1'.
                    complete_name = f"{rack_str}-{str(pos_int).zfill(2)}-N{nivel_str}-P{puesto_str}"

                    existing = self.env['stock.location'].search([
                        ('complete_name', '=', complete_name),
                        ('warehouse_id', '=', self.warehouse_id.id),
                    ], limit=1)

                    if existing:
                        existing.write({
                            'is_map_location': True,
                            'map_rack': rack_str,
                            'map_position': pos_int,
                            'map_level': nivel_str,
                            'map_slot': puesto_str,
                        })
                    else:
                        self.env['stock.location'].create({
                            'name': complete_name,
                            'complete_name': complete_name,
                            'warehouse_id': self.warehouse_id.id,
                            'usage': 'internal',
                            'location_id': self.warehouse_id.view_location_id.id,
                            'is_map_location': True,
                            'map_rack': rack_str,
                            'map_position': pos_int,
                            'map_level': nivel_str,
                            'map_slot': puesto_str,
                        })
                    success += 1
                except Exception as e:
                    errors += 1
                    log_lines.append(f"Fila {row_idx}: Error - {e}")

            wb.close()
        except Exception as e:
            log_lines.append(f"Error al abrir archivo: {e}")
            errors += 1

        log_lines.insert(0, f"Importación finalizada: {success} exitosas, {errors} errores.")
        self.log = '\n'.join(log_lines)
        self.state = 'result'
        return self._reload()

    def _reload(self):
        """Recarga el formulario en modo resultado (muestra el log)."""
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'warehouse.map.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
